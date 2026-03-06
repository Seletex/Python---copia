import os
import sys
import pandas as pd

# Asegurar que los módulos en la carpeta Actividades estén en sys.path
ROOT = os.path.abspath(os.path.join(os.getcwd()))
ACT_DIR = os.path.join(ROOT, 'Actividades')
if ACT_DIR not in sys.path:
    sys.path.insert(0, ACT_DIR)

# Importar utilidades de la app
from database_setup import init_db
from database import guardar_configuracion_usuario
from export_service import generar_informe_template

from openpyxl import load_workbook


def main():
    print('Iniciando prueba de generación de informe...')
    init_db()

    # Guardar datos de contrato bajo admin
    datos = {
        'nro': 'C-TEST-001',
        'objeto': 'Contrato de prueba para generación',
        'nombre': 'EMPRESA PRUEBA S.A.S.',
        'cedula': '900123456-7',
        'supervisor': 'SUPERVISOR DEMO'
    }
    ok = guardar_configuracion_usuario('admin', {'datos_contrato': datos})
    print('Guardar datos contrato:', ok)

    # Crear DataFrame de ejemplo
    df = pd.DataFrame([{
        'USUARIO': 'usuario1',
        'FECHA': '2026-03-03 10:00:00',
        'TIPO DE ACTIVIDAD': 'INSTALACIÓN',
        'DEPENDENCIA': 'OFICINA',
        'SOLICITANTE': 'Solicitante X',
        'TIPO DE SOLICITUD': 'MANTENIMIENTO PREVENTIVO',
        'MEDIO DE SOLICITUD': 'INTRANET',
        'CUMPLIDO': 'Sí',
        'FECHA ATENCIÓN': '2026-03-04',
        'OBSERVACIONES': 'Prueba automática'
    }])

    out = os.path.join(ACT_DIR, 'test_informe_generado.xlsx')
    if os.path.exists(out):
        try:
            os.remove(out)
        except:
            pass

    resultado = generar_informe_template(df, out, contrato_data=None)
    print('generar_informe_template returned:', resultado)

    if not resultado or not os.path.exists(out):
        print('ERROR: no se generó el archivo de salida')
        return

    wb = load_workbook(out)
    ws = wb.active

    print('Celda (2,3):', ws.cell(row=2, column=3).value)
    print('Celda (3,3):', ws.cell(row=3, column=3).value)
    print('Celda (4,3):', ws.cell(row=4, column=3).value)
    print('Celda (6,3):', ws.cell(row=6, column=3).value)

if __name__ == '__main__':
    main()
