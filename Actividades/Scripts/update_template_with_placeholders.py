"""
Script para actualizar la plantilla Excel con placeholders legibles.
Lee la plantilla actual y reemplaza celdas fijas con marcadores.
"""
import os
import sys
import shutil
from openpyxl import load_workbook

ROOT = os.path.abspath(os.path.join(os.getcwd()))
ACT_DIR = os.path.join(ROOT, 'Actividades')
sys.path.insert(0, ACT_DIR)

from config import TEMPLATE_EXCEL, DEFAULT_DATA_DIR

def update_template_with_placeholders():
    """Actualiza la plantilla reemplazando celdas fijas con placeholders."""
    
    if not os.path.exists(TEMPLATE_EXCEL):
        print(f"ERROR: Plantilla no encontrada en {TEMPLATE_EXCEL}")
        return False
    
    # Hacer backup de la plantilla original
    backup_path = TEMPLATE_EXCEL.replace('.xlsx', '_backup.xlsx')
    try:
        shutil.copy2(TEMPLATE_EXCEL, backup_path)
        print(f"✓ Backup creado: {backup_path}")
    except Exception as e:
        print(f"⚠ No se pudo crear backup: {e}")
    
    try:
        # Cargar la plantilla
        wb = load_workbook(TEMPLATE_EXCEL)
        ws = wb.active
        
        # Reemplazar celdas con placeholders
        placeholders = {
            (2, 3): '{{NRO_CONTRATO}}',
            (3, 3): '{{OBJETO}}',
            (4, 3): '{{NOMBRE_CONTRATISTA}}',
            (5, 3): '{{CEDULA}}',
            (6, 3): '{{RANGO_FECHAS}}'
        }
        
        for (row, col), placeholder in placeholders.items():
            cell = ws.cell(row=row, column=col)
            print(f"Celda ({row},{col}): '{cell.value}' → '{placeholder}'")
            cell.value = placeholder
        
        # Guardar la plantilla actualizada
        wb.save(TEMPLATE_EXCEL)
        print(f"✓ Plantilla actualizada: {TEMPLATE_EXCEL}")
        return True
        
    except Exception as e:
        print(f"ERROR: No se pudo actualizar la plantilla: {e}")
        return False

if __name__ == '__main__':
    print(f"Directorio de datos: {DEFAULT_DATA_DIR}")
    print(f"Plantilla: {TEMPLATE_EXCEL}\n")
    success = update_template_with_placeholders()
    if success:
        print("\n✓ Plantilla actualizada con placeholders.")
    else:
        print("\n✗ Error al actualizar la plantilla.")
