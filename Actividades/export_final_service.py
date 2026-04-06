from config import logger, TEMPLATE_INFORME_FINAL
from utils import medir_tiempo
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

@medir_tiempo
def generar_informe_final_resumen(df, output_path, contrato_data=None, usuario=None):
    """
    Genera el Informe Final Concentrado cargando la plantilla original.
    """
    try:
        def _escribir_seguro(ws, r, c, val):
            """Escribe en una celda, buscando la maestra si está combinada"""
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                for m_range in ws.merged_cells.ranges:
                    if cell.coordinate in m_range:
                        ws.cell(row=m_range.min_row, column=m_range.min_col, value=val)
                        return
            cell.value = val

        logger.info(f"Generando Informe Final Concentrado. Registros iniciales: {len(df)}")

        if df.empty:
            logger.warning("Generar Informe Final: DataFrame vacío")
            return False

        # --- FILTRO POR USUARIO ---
        if usuario:
            df = df[df['USUARIO'] == usuario].copy()
            logger.info(f"Reporte Final: Filtrado para usuario '{usuario}'. Registros: {len(df)}")
        
        if df.empty:
            logger.warning(f"Reporte Final: No hay registros para el usuario '{usuario}' después de filtrar.")
            return False

        # Obtener datos de contrato si faltan
        if contrato_data is None:
            try:
                from database import obtener_configuracion_usuario
                cfg = obtener_configuracion_usuario('admin') or {}
                contrato_data = cfg.get('datos_contrato', {}) if isinstance(cfg, dict) else {}
            except Exception:
                contrato_data = {}

        if not os.path.exists(TEMPLATE_INFORME_FINAL):
            logger.error(f"No se encontró la plantilla final en: {TEMPLATE_INFORME_FINAL}")
            # Fallback al generador dinámico si no hay plantilla (opcional, pero mejor avisar)
            return False

        wb = load_workbook(TEMPLATE_INFORME_FINAL)
        ws = wb.active

        # Datos del contrato a celdas fijas
        nro = (contrato_data.get('nro') if contrato_data else '') or ''
        objeto = (contrato_data.get('objeto') if contrato_data else '') or ''
        nombre = (contrato_data.get('nombre') if contrato_data else '') or (usuario.upper() if usuario else '')
        cedula = (contrato_data.get('cedula') if contrato_data else '') or ''
        supervisor = (contrato_data.get('supervisor') if contrato_data else '') or ''

        # Rango fechas
        rango_fechas = ''
        if 'FECHA' in df.columns:
            fechas_dt = pd.to_datetime(df['FECHA'], errors='coerce').dropna()
            if not fechas_dt.empty:
                rango_fechas = f"{fechas_dt.min().strftime('%d/%m/%Y')} al {fechas_dt.max().strftime('%d/%m/%Y')}"

        # Llenado de metadatos (Basado en escaneo de plantilla real con función segura)
        _escribir_seguro(ws, 2, 3, str(nro).upper())
        _escribir_seguro(ws, 3, 3, str(objeto).upper())
        _escribir_seguro(ws, 4, 3, str(nombre).upper())
        _escribir_seguro(ws, 5, 3, str(cedula))
        _escribir_seguro(ws, 6, 3, rango_fechas)
        _escribir_seguro(ws, 8, 2, datetime.now().strftime('%d/%m/%Y')) # Fecha informe

        # 1. Agrupar y contar actividades
        if 'TIPO DE ACTIVIDAD' in df.columns:
            # Normalizar para evitar duplicados por espacios
            df_norm = df.copy()
            df_norm['TIPO DE ACTIVIDAD'] = df_norm['TIPO DE ACTIVIDAD'].fillna('OTRO').str.strip()
            resumen = df_norm['TIPO DE ACTIVIDAD'].value_counts().reset_index()
            resumen.columns = ['Actividad', 'Cantidad']
        else:
            resumen = pd.DataFrame(columns=['Actividad', 'Cantidad'])

        # Estilo para bordes
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Inserción de filas para actividades (Iniciando en fila 8)
        start_row = 8
        if not resumen.empty:
            # Insertamos el espacio necesario para las actividades
            ws.insert_rows(start_row, amount=len(resumen))
            
            for i, (_, r) in enumerate(resumen.iterrows()):
                curr_row = start_row + i
                # Columna A (1): Actividad
                cell_act = ws.cell(row=curr_row, column=1, value=r['Actividad'])
                cell_act.border = thin_border
                cell_act.alignment = Alignment(wrap_text=True)
                ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=7)
                
                # Columna H (8): Cantidad
                cell_cant = ws.cell(row=curr_row, column=8, value=int(r['Cantidad']))
                cell_cant.border = thin_border
                cell_cant.alignment = center_align

        # Actualizar Firmas (Bajarán según las filas insertadas)
        # Buscamos las celdas de firma después de la tabla
        final_idx = start_row + len(resumen) + 1
        ws.cell(row=final_idx+1, column=2, value=str(nombre).upper()) # Elaborado por
        ws.cell(row=final_idx+3, column=2, value=str(supervisor).upper()) # Vo Bo

        wb.save(output_path)
        logger.info(f"Informe Final generado usando plantilla en: {output_path}")
        return True

    except Exception as e:
        logger.error(f"Error generando informe final: {e}")
        return False
