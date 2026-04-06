"""
Servicio de exportación y generación de reportes.
Separado de database.py para responsabilidad única.
"""

import os
import copy
import pandas as pd
import unicodedata
from datetime import datetime
from config import TEMPLATE_EXCEL, logger
from database import cargar_registros
from utils import medir_tiempo


@medir_tiempo
def exportar_registros_filtrados(fecha_inicio=None, fecha_fin=None, usuario=None, actividad=None):
    """Exporta registros filtrados. Retorna (DataFrame, dict_estadísticas)"""
    try:
        df = cargar_registros(usuario)
        if df.empty:
            return pd.DataFrame(), {}
        
        # Parsear fechas
        if 'FECHA' in df.columns:
            df['FECHA'] = pd.to_datetime(df['FECHA'], errors='coerce')
        
        # Aplicar filtros
        if fecha_inicio:
            df = df[df['FECHA'] >= pd.to_datetime(fecha_inicio)]
        if fecha_fin:
            df = df[df['FECHA'] <= pd.to_datetime(fecha_fin)]
        if actividad and actividad != 'Todas' and 'TIPO DE ACTIVIDAD' in df.columns:
            df = df[df['TIPO DE ACTIVIDAD'] == actividad]
        
        # Agrupamiento y ordenamiento solicitado
        if not df.empty:
            sort_cols = []
            if 'TIPO DE ACTIVIDAD' in df.columns: sort_cols.append('TIPO DE ACTIVIDAD')
            if 'FECHA' in df.columns: sort_cols.append('FECHA')
            if sort_cols:
                df = df.sort_values(by=sort_cols)
        
        stats = _calcular_estadisticas(df)
        return df, stats
    except Exception as e:
        logger.error(f"Error exportando registros: {e}")
        return pd.DataFrame(), {}


def _calcular_estadisticas(df):
    """Calcula estadísticas básicas de un DataFrame"""
    if df.empty:
        return {}
    
    stats = {
        'total_registros': len(df),
        'fecha_inicio': _format_fecha(df, 'min'),
        'fecha_fin': _format_fecha(df, 'max'),
    }
    
    # Conteos por columna
    for col, key in [('TIPO DE ACTIVIDAD', 'conteo_por_actividad'),
                     ('TIPO DE SOLICITUD', 'conteo_por_solicitud'),
                     ('MEDIO DE SOLICITUD', 'conteo_por_medio')]:
        if col in df.columns:
            counts = df[col].value_counts().to_dict()
            stats[key] = {str(k): int(v) for k, v in counts.items()}
    
    return stats


def _format_fecha(df, func):
    """Formatea fecha min/max de un DataFrame"""
    if 'FECHA' not in df.columns or df['FECHA'].empty:
        return 'N/A'
    try:
        val = getattr(df['FECHA'], func)()
        return val.strftime('%Y-%m-%d') if pd.notna(val) else 'N/A'
    except Exception:
        return 'N/A'


@medir_tiempo
def obtener_estadisticas_exportacion(usuario=None, fecha_inicio=None, fecha_fin=None, **kwargs):
    """Obtiene estadísticas generales y datos para gráficos con soporte de filtros"""
    empty_result = {
        'fecha_min': 'N/A', 'fecha_max': 'N/A',
        'total_registros': 0, 'total_tipos_actividad': 0,
        'ultima_exportacion': 'Nunca',
        'chart_actividades': {'labels': [], 'data': []},
        'chart_cumplimiento': {'labels': [], 'data': []},
        'chart_linea': {'labels': [], 'data': []},
        'usuarios': []
    }
    
    try:
        df, _ = exportar_registros_filtrados(
            fecha_inicio=fecha_inicio, 
            fecha_fin=fecha_fin, 
            usuario=usuario,
            actividad=kwargs.get('actividad')
        )
        if df.empty:
            return empty_result
        
        # Parseo de fechas (exportar_registros_filtrados ya hace parte del proceso)
        if 'FECHA' in df.columns:
            df['FECHA_DT'] = pd.to_datetime(df['FECHA'], errors='coerce')
            df = df.dropna(subset=['FECHA_DT'])
        
        if df.empty:
            return empty_result
        
        # Gráfico: Actividades (Todas, según petición del usuario)
        counts_act = df['TIPO DE ACTIVIDAD'].value_counts()
        chart_actividades = {
            'labels': counts_act.index.tolist(),
            'data': counts_act.values.tolist()
        }
        
        # Gráfico: Cumplimiento
        cumplimiento = df['CUMPLIDO'].value_counts() if 'CUMPLIDO' in df.columns else pd.Series()
        chart_cumplimiento = {
            'labels': cumplimiento.index.tolist(),
            'data': cumplimiento.values.tolist()
        }
        
        # Gráfico: Línea temporal
        df_sorted = df.sort_values('FECHA_DT')
        linea = df_sorted['FECHA_DT'].dt.date.value_counts().sort_index()
        # Si hay demasiados días, mostrar los últimos 90 para no saturar
        if len(linea) > 90:
            linea = linea.tail(90)
            
        chart_linea = {
            'labels': [d.strftime('%d/%m') for d in linea.index],
            'data': linea.values.tolist()
        }
        
        # Estadística por usuario
        user_stats = []
        if 'USUARIO' in df.columns:
            for user, group in df.groupby('USUARIO'):
                total_user = len(group)
                cumplidos = len(group[group['CUMPLIDO'] == 'Sí'])
                porcentaje = f"{(cumplidos/total_user)*100:.1f}%" if total_user > 0 else "0%"
                ultima = group['FECHA_DT'].max().strftime('%Y-%m-%d %H:%M') if not group.empty else "N/A"
                user_stats.append({
                    'usuario': user,
                    'total': total_user,
                    'cumplimiento': porcentaje,
                    'ultima': ultima
                })
        
        return {
            'fecha_min': df['FECHA_DT'].min().strftime('%Y-%m-%d'),
            'fecha_max': df['FECHA_DT'].max().strftime('%Y-%m-%d'),
            'total_registros': len(df),
            'total_tipos_actividad': df['TIPO DE ACTIVIDAD'].nunique() if 'TIPO DE ACTIVIDAD' in df.columns else 0,
            'ultima_exportacion': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'chart_actividades': chart_actividades,
            'chart_cumplimiento': chart_cumplimiento,
            'chart_linea': chart_linea,
            'usuarios': user_stats
        }
    except Exception as e:
        logger.error(f"Error obteniendo estadísticas: {e}")
        return empty_result


@medir_tiempo
def generar_reporte_excel(df, estadisticas, output_path):
    """Genera un archivo Excel con datos + estadísticas"""
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Registros', index=False)
            
            stats_df = pd.DataFrame([
                ['Total de registros', estadisticas.get('total_registros', 0)],
                ['Fecha inicio', estadisticas.get('fecha_inicio', 'N/A')],
                ['Fecha fin', estadisticas.get('fecha_fin', 'N/A')]
            ], columns=['Métrica', 'Valor'])
            stats_df.to_excel(writer, sheet_name='Estadísticas', index=False)
            
            if 'conteo_por_actividad' in estadisticas:
                act_df = pd.DataFrame(
                    estadisticas['conteo_por_actividad'].items(),
                    columns=['Actividad', 'Cantidad']
                )
                act_df.to_excel(writer, sheet_name='Estadísticas', startrow=5, index=False)
        
        return True
    except Exception as e:
        logger.error(f"Error generando reporte Excel: {e}")
        return False


@medir_tiempo
def generar_informe_template(df, output_path, contrato_data=None):
    """Genera informe usando la plantilla Excel original para preservar el formato profesional"""
    try:
        import os
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.cell.cell import MergedCell
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from config import logger, TEMPLATE_EXCEL

        def _escribir_seguro(ws, r, c, val):
            """Escribe en una celda, buscando la maestra si está combinada"""
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                for m_range in ws.merged_cells.ranges:
                    if cell.coordinate in m_range:
                        ws.cell(row=m_range.min_row, column=m_range.min_col, value=val)
                        return
            cell.value = val

        if not os.path.exists(TEMPLATE_EXCEL):
            logger.error(f"No se encontró la plantilla detallada en: {TEMPLATE_EXCEL}")
            return False

        if contrato_data is None:
            try:
                from database import obtener_configuracion_usuario
                cfg = obtener_configuracion_usuario('admin') or {}
                contrato_data = cfg.get('datos_contrato', {}) if isinstance(cfg, dict) else {}
            except Exception:
                contrato_data = {}

        wb = load_workbook(TEMPLATE_EXCEL)
        ws = wb.active

        # Datos del contrato a celdas fijas (Basado en análisis de plantilla detallada)
        nro = (contrato_data.get('nro') if contrato_data else '') or ''
        objeto = (contrato_data.get('objeto') if contrato_data else '') or ''
        nombre = (contrato_data.get('nombre') if contrato_data else '') or ''
        cedula = (contrato_data.get('cedula') if contrato_data else '') or ''
        supervisor = (contrato_data.get('supervisor') if contrato_data else '') or ''

        # Rango fechas
        rango_fechas = ''
        if not df.empty and 'FECHA' in df.columns:
            fechas_dt = pd.to_datetime(df['FECHA'], errors='coerce').dropna()
            if not fechas_dt.empty:
                rango_fechas = f"{fechas_dt.min().strftime('%d/%m/%Y')} al {fechas_dt.max().strftime('%d/%m/%Y')}"

        # Llenado de metadatos (Uso de función segura para celdas combinadas)
        _escribir_seguro(ws, 2, 3, str(nro).upper())
        _escribir_seguro(ws, 3, 3, str(objeto).upper())
        _escribir_seguro(ws, 4, 3, str(nombre).upper())
        _escribir_seguro(ws, 5, 3, str(cedula))
        _escribir_seguro(ws, 6, 3, rango_fechas)

        # Estilos base para las celdas de datos
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Iniciar datos en la fila 8 (La fila 7 tiene los encabezados originales)
        row_idx = 8
        
        # Agrupar por tipo de actividad para facilitar lectura
        df_reporte = df.copy()
        if 'TIPO DE ACTIVIDAD' in df_reporte.columns:
            df_reporte = df_reporte.sort_values(by=['TIPO DE ACTIVIDAD', 'FECHA'])

        ultima_actividad = None
        subtotal_actividad = 0
        total_general = 0

        def _escribir_subtotal(idx, act, count):
            ws.cell(row=idx, column=1, value=f"SUBTOTAL {act}:").font = Font(bold=True)
            ws.cell(row=idx, column=2, value=count).font = Font(bold=True)
            ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=9)
            for i in range(1, 10):
                ws.cell(row=idx, column=i).border = thin_border
                ws.cell(row=idx, column=i).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Columnas mapeadas: TIPO DE ACTIVIDAD(1), FECHA(2), DEPENDENCIA(3), SOLICITANTE(4), 
        # TIPO DE SOLICITUD(5), MEDIO DE SOL.(6), CUMPLIDO(7), FECHA ATENCIÓN(8), OBSERVACIONES(9)
        columnas_mapeo = [
            'TIPO DE ACTIVIDAD', 'FECHA ATENCIÓN', 'DEPENDENCIA', 'SOLICITANTE',
            'TIPO DE SOLICITUD', 'MEDIO DE SOLICITUD', 'CUMPLIDO', 'FECHA ATENCIÓN', 'OBSERVACIONES'
        ]

        for _, r in df_reporte.iterrows():
            act_actual = r.get('TIPO DE ACTIVIDAD', 'SIN TIPO')
            
            # Escribir registros
            for col_idx, col_name in enumerate(columnas_mapeo, 1):
                val = r.get(col_name, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else '')
                cell.border = thin_border
                cell.alignment = left_align if col_idx in [1, 3, 4, 9] else center_align
                
            total_general += 1
            row_idx += 1

        # TOTAL GENERAL
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="TOTAL GENERAL DE REGISTROS:").font = Font(bold=True, size=12)
        ws.cell(row=row_idx, column=2, value=total_general).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=9)
        for i in range(1, 10):
            c = ws.cell(row=row_idx, column=i)
            c.border = thin_border
            c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
        # ------------------
        # SECCIÓN FIRMAS
        # ------------------
        row_idx += 4
        ws.cell(row=row_idx, column=2, value="_______________________________").alignment = center_align
        ws.cell(row=row_idx, column=6, value="_______________________________").alignment = center_align
        row_idx += 1
        ws.cell(row=row_idx, column=2, value="FIRMA CONTRATISTA").font = Font(bold=True)
        ws.cell(row=row_idx, column=2).alignment = center_align
        ws.cell(row=row_idx, column=6, value="FIRMA SUPERVISOR").font = Font(bold=True)
        ws.cell(row=row_idx, column=6).alignment = center_align
        
        row_idx += 1
        ws.cell(row=row_idx, column=2, value=str(nombre).upper()).alignment = center_align
        ws.cell(row=row_idx, column=6, value=str(supervisor).upper()).alignment = center_align
        
        row_idx += 1
        ws.cell(row=row_idx, column=2, value=str(cedula)).alignment = center_align
        
        wb.save(output_path)
        logger.info(f"Informe generado limpiamente en: {output_path}")
        return True

    except Exception as e:
        from config import logger
        logger.exception(f"Excepción en generar_informe_template: {e}")
        return False


@medir_tiempo
def analizar_plantilla_contrato(df=None, contrato_data=None, output_json_path=None):
    """Obsoleta: Retorna vacío ya que no se depende de plantilla física."""
    return {}
